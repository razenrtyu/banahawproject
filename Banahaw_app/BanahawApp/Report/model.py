import datetime
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Alignment
from openpyxl.cell import Cell
import os
from flask import request
from sqlalchemy import and_
from BanahawApp import Session,Mini_func, app
from BanahawApp.table import T_Transaction, T_Member00, T_Attendants01, T_Attendants


class Attendant_report(object):

	def __init__(self, **kwargs):
		self.__session = Session()
		self.__args = kwargs
		self.__retval = None
		self.__tempdict = dict()
		self.__total = 0

	def get_reports_data(self):
		"""."""
		ds = self.__args.get('from',None)
		de = self.__args.get('to',None)
		by_id = self.__args.get('attendantid', None)

		if all([ds,de,by_id]):
			self.__tempdict['data'] = list()
			transactionstotal, transsales = self.__get_transactions(ds, de, by_id)
			memberstotal, memsales = self.__get_members(ds, de, by_id)
			upgradestotal, upgradessales = self.__get_members_upg(ds, de, by_id)

		if self.__tempdict:
			self.__retval = self.__tempdict
			self.__retval['total_on_service'] = transactionstotal
			self.__retval['total_on_membership'] = memberstotal + upgradestotal
			self.__retval['total_sales'] = transsales + memsales + upgradessales

		return self.__retval

	def __get_transactions(self, ds, de, attid):
		search_filter = list()

		search_filter.append(getattr(T_Transaction, 'datecreated').between(ds,de))
		search_filter.append(getattr(T_Transaction, 'attendantid') == attid)
		search_filter.append(getattr(T_Transaction, 'active') == 0)

		result = self.__session.query(T_Transaction).filter(
			and_(*search_filter)).all()

		total = 0
		total2 = 0

		for data in result:
			key = data.datecreated.strftime('%B-%d-%Y')

			# if key not in self.__tempdict:
			# 	self.__tempdict[key] = list()

			comm = self.__get_commision_on_services(data.service,data.service_price,
											 		data.transaction_type,data.add_ons_price)

			total += comm
			total2 += data.total_amount

			data_dict = dict()
			data_dict = {
				'datecreated': key,
				'clientname': data.client_name,
				'services': data.service + ', ' + data.add_ons,
				'timespent': '{:02d}H:{:02d}M'.format(*divmod(data.time_spent, 60)),
				'amountpaid': data.total_amount,
				'commision_on_service': comm
			}

			self.__tempdict['data'].append(data_dict)

		return total, total2

	def __get_commision_on_services(self, service, serviceprice, trantype, add_ons_price):
		ignore_list = ['5-in-1 Signature Massage', 'Relaxing Swedish Massage']
		retval = 0

		servicecomm = 0
		if service not in ignore_list:
			if serviceprice is None:
				serviceprice = 0

			servicecomm = int(serviceprice) * 0.10

		if not add_ons_price:
			return servicecomm

		addonsprices = add_ons_price.split(',')
		addonscomm = 0

		if trantype in ['Walk-In','Non-Member']:		
			for price in addonsprices:
				if int(price) <= 299:
					addonscomm += 25
				elif int(price) >= 300:
					addonscomm += 50

		elif trantype == 'Member':
			for price in addonsprices:
				if int(price) <= 149:
					addonscomm += 25
				elif int(price) >= 150:
					addonscomm += 50

		retval = addonscomm + servicecomm

		return retval

	def __get_members(self, ds, de, attid):
		retval = list()
		search_filter = list()

		search_filter.append(getattr(T_Member00, 'datecreated').between(ds, de))
		search_filter.append(getattr(T_Member00, 'attendantid') == attid)

		result = self.__session.query(T_Member00).filter(and_(*search_filter)).all()

		total = 0
		total2 = 0

		for data in result:
			key = data.datecreated.strftime('%B-%d-%Y')

			# if key not in self.__tempdict:
			# 	self.__tempdict[key] = list()

			if data.upgraded:
				total += 25
				total2 += 300

				data_dict = dict()
				data_dict = {
					'clientname': data.name,
					'services': 'Membership Sold - ' + 'Personalized',
					'amountpaid': 300,
					'commision_on_service': 25,
					'timespent': 'N/A',
					'datecreated':key
				}
			else:

				comm = 25 if data.membertype == 'Personalized' else 50

				total += comm
				total2 += data.membershipcost

				data_dict = dict()
				data_dict = {
					'clientname': data.name,
					'services': 'Membership Sold - ' + data.membertype,
					'amountpaid': data.membershipcost,
					'commision_on_service': comm,
					'timespent': 'N/A',
					'datecreated':key
				}

			self.__tempdict['data'].append(data_dict)

		return total, total2

	def __get_members_upg(self, ds, de, attid):
		retval = list()
		search_filter = list()

		search_filter.append(getattr(T_Member00, 'upgraded').between(ds, de))
		search_filter.append(getattr(T_Member00, 'upgraded_by') == attid)

		result = self.__session.query(T_Member00).filter(and_(*search_filter)).all()

		total = 0
		total2 = 0

		for data in result:
			key = data.upgraded.strftime('%B-%d-%Y')

			# if key not in self.__tempdict:
			# 	self.__tempdict[key] = list()

			total += 25
			total2 += 300

			data_dict = dict()
			data_dict = {
				'clientname': data.name,
				'services': 'Membership Sold - ' + data.membertype + '[UPGRADED]',
				'amountpaid': 300,
				'commision_on_service': 25,
				'timespent': 'N/A',
				'datecreated':key
			}

			self.__tempdict['data'].append(data_dict)

		return total, total2


class Summary_report(object):
	def __init__(self, **kwargs):
		self.__session = Session()
		self.__args = kwargs
		self.__retval = dict()

	def get_reports_data(self):
		retval = dict()

		ds = self.__args.get('from',None)
		de = self.__args.get('to',None)

		self.__get_attendants(ds, de)
		self.__get_rawtime(ds, de)
		self.__get_transactions(ds, de)
		self.__get_members(ds, de)
		self.__get_members_upg(ds, de)


		if self.__retval:
			total_allowance = 0
			total_comm = 0
			total_incentives = 0

			for attid, data in self.__retval.items():
				self.__retval[attid]['total_per_att'] = data.get('service_comm') + data.get('mem_incentive')
				total_allowance += data.get('allowance', 0)
				total_comm += data.get('service_comm', 0)
				total_incentives += data.get('mem_incentive')

			retval['data'] = self.__retval
			retval['totals'] = {
				'TOTAL ALLOWANCE': total_allowance,
				'TOTAL COMMISION': total_comm,
				'TOTAL INCENTIVES ON MEMBERSHIP': total_incentives,
				'TOTAL GROSS SALES FOR DAY': 0,
				'TOTAL NET SALES': 0
			}

		dateds = datetime.datetime.strptime(ds, '%Y-%m-%d')
		datede = datetime.datetime.strptime(de, '%Y-%m-%d')
		delta = datetime.timedelta(days=1)
		headers = ['Name of Employee']

		while dateds <= datede:
			key = dateds.strftime('%B-%d-%Y')
			headers.append(key)
			dateds += delta

		headers.append('Allowance')
		headers.append('Commision on Service')
		headers.append('Incentive on Membership')
		headers.append('Total per Attendant')


		self.__make_excelfile(retval, headers)

		return retval

	def __make_excelfile(self, jsondata, headers):

		temp_list = list()
		for key, val in jsondata['data'].items():
			temp_dict = dict()
			for key2, val2 in val.items():
				if key2 == 'attendant_name':
					temp_dict['Name of Employee'] = val2
				elif key2 == 'mem_incentive':
					temp_dict['Incentive on Membership'] = val2
				elif key2 == 'service_comm':
					temp_dict['Commision on Service'] = val2
				elif key2 == 'total_per_att':
					temp_dict['Total per Attendant'] = val2
				elif key2 =='allowance':
					temp_dict['Allowance'] = val2
				else:
					temp_dict[key2] = val2
			temp_list.append(temp_dict)

		filename = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
		folderpath = os.path.join(os.getcwd(), 'Reports')
		if not os.path.exists(folderpath):
			os.makedirs(folderpath)

		wb = openpyxl.Workbook()
		sheet = wb.active
		sheet.title = 'Sales Reports Summary'

		for row, data in enumerate(temp_list):
			excel_row = row + 1
			for index, key in enumerate(headers):
				if excel_row == 1:
					x = sheet.cell(row=excel_row, column=index + 1)
					x.value = key
					xcell = sheet[x.coordinate]
					ft = Font(bold=True)
					al = Alignment(horizontal='center', vertical='center')
					xcell.font = ft
					xcell.alignment = al
					sheet.row_dimensions[int(x.coordinate[-1:])].height = 45
					sheet.column_dimensions[str(x.coordinate[:-1])].width = 25
				else:
					x = sheet.cell(row=excel_row, column=index + 1)
					xcell = sheet[x.coordinate]
					al = Alignment(horizontal='center', vertical='center')
					xcell.alignment = al
					x.value = data[key]

		rownum = len(temp_list) + 6
		colorcoding = {
			'TOTAL ALLOWANCE': 'ff8080',
			'TOTAL COMMISION': 'ffd6cc',
			'TOTAL INCENTIVES ON MEMBERSHIP': 'ccccff',
			'TOTAL GROSS SALES FOR DAY': 'ccffdd',
			'TOTAL NET SALES': 'e6ccb3'
		}
		for key,val in jsondata['totals'].items():
			ft = Font(italic=True)
			al = Alignment(horizontal='center', vertical='center')
			fl = PatternFill(start_color=colorcoding[key], end_color=colorcoding[key], fill_type='solid')
			mergecells = 'A{0}:B{0}'.format(rownum)
			sheet.merge_cells(mergecells)

			sheet['A{0}'.format(rownum)].font = ft
			sheet['A{0}'.format(rownum)].alignment = al
			sheet['A{0}'.format(rownum)].fill = fl
			sheet['A{0}'.format(rownum)] = key

			sheet['C{0}'.format(rownum)].alignment = al
			sheet['C{0}'.format(rownum)].fill = fl
			sheet['C{0}'.format(rownum)] = val

			rownum += 1

		wb.save(os.path.join(folderpath, 'Sales_report_' + filename + '.xlsx'))




	def __get_attendants(self, ds, de):
		result = self.__session.query(T_Attendants).all()

		for data in result:
			self.__retval[data.attendantid] = {
				'attendant_name': data.attendant_name,
				'service_comm': 0,
				'allowance':0,
				'mem_incentive': 0
			}

			datestart = datetime.datetime.strptime(ds, '%Y-%m-%d')
			dateend = datetime.datetime.strptime(de, '%Y-%m-%d')
			delta = datetime.timedelta(days=1)

			total_allwowance = 0
			while datestart <= dateend:
				key = datestart.strftime('%B-%d-%Y')
				self.__retval[data.attendantid][key] = 'No Time In - No Time Out'
				total_allwowance += data.allowance
				datestart += delta

			# assign total allowance
			self.__retval[data.attendantid]['allowance'] = total_allwowance

	def __get_rawtime(self, ds, de):
		search_filter = list()

		search_filter.append(getattr(T_Attendants01, 'trandate').between(ds,de))

		result = self.__session.query(T_Attendants01).filter(and_(*search_filter)).all()
		for data in result:
			if not data.timein:
				timein = 'No Time In'
			else:
				timein = data.timein

			if not data.timeout:
				timeout = 'No Time Out'
			else:
				timeout = data.timeout

			key = data.trandate.strftime('%B-%d-%Y')

			self.__retval[data.attendantid][key] = timein + ' - ' + timeout

	def __get_transactions(self, ds, de):
		search_filter = list()

		search_filter.append(getattr(T_Transaction, 'datecreated').between(ds,de))
		search_filter.append(getattr(T_Transaction, 'active') == 0)

		result = self.__session.query(T_Transaction).filter(
			and_(*search_filter)).order_by(T_Transaction.attendantid).all()

		total = 0

		for data in result:
			if data.attendantid not in self.__retval:
				total = 0
				total2 = 0

			comm = self.__get_commision_on_services(data.service,data.service_price,
											 		data.transaction_type,data.add_ons_price)
			total += comm

			# assign total commision
			self.__retval[data.attendantid]['service_comm'] = total


	def __get_commision_on_services(self, service, serviceprice, trantype, add_ons_price):
		ignore_list = ['5-in-1 Signature Massage', 'Relaxing Swedish Massage']
		retval = 0

		servicecomm = 0
		if service not in ignore_list:
			if serviceprice is None:
				serviceprice = 0

			servicecomm = int(serviceprice) * 0.10
		if not add_ons_price:
			return servicecomm

		addonsprices = add_ons_price.split(',')
		addonscomm = 0

		if trantype in ['Walk-In','Non-Member']:		
			for price in addonsprices:
				if int(price) <= 299:
					addonscomm += 25
				elif int(price) >= 300:
					addonscomm += 50

		elif trantype == 'Member':
			for price in addonsprices:
				if int(price) <= 149:
					addonscomm += 25
				elif int(price) >= 150:
					addonscomm += 50

		retval = addonscomm + servicecomm

		return retval

	def __get_members(self, ds, de):
		retval = list()
		search_filter = list()

		search_filter.append(getattr(T_Member00, 'datecreated').between(ds, de))

		result = self.__session.query(T_Member00).filter(and_(*search_filter)).all()

		total = 0

		for data in result:
			if data.upgraded:
				total += 25

			else:
				comm = 25 if data.membertype == 'Personalized' else 50

				total += comm

			self.__retval[data.attendantid]['mem_incentive'] = total

	def __get_members_upg(self, ds, de):
		search_filter = list()

		search_filter.append(getattr(T_Member00, 'upgraded').between(ds, de))
		search_filter.append(getattr(T_Member00, 'upgraded_by') != None)

		result = self.__session.query(T_Member00).filter(and_(*search_filter)).all()

		total = 0
		total2 = 0

		for data in result:
			total += 25

			self.__retval[data.attendantid]['mem_incentive'] += total

@app.route("/download-reports", methods=['POST'])
def download_file():
	data = request.get_json()
	name = datetime.datetime.now().strftime('%B-%d-%Y') + '_Summary_Report'

	return excel.make_response_from_records(x, "xlsx",200,name)
